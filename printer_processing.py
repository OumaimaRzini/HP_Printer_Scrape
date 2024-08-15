import openpyxl

def calculate_difference(excel_file, sheet_name):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        
        # Select the worksheet
        sheet = wb[sheet_name]
        
        # Create a new worksheet for differences
        diff_sheet = wb.create_sheet(title="Printer Daily Usage")
        
        # Copy the original data to the differences sheet
        for row in sheet.iter_rows(values_only=True):
            diff_sheet.append(row)
        
        # Dictionary to store the previous row for each printer ID
        prev_rows = {}
        
        # Iterate over rows starting from the second row (assuming headers are in the first row)
        for index, row in enumerate(diff_sheet.iter_rows(min_row=2, values_only=True), start=2):
            printer_id = row[0]  # Assuming printer ID is in the first column
            
            # If printer ID is not already in the dictionary, add it with the current row
            if printer_id not in prev_rows:
                prev_rows[printer_id] = row
            else:
                # Calculate the differences from the previous row
                for i, col in enumerate(["F", "G"], start=5):  
                    try:
                        current_value = int(row[i])
                        prev_row_value = int(prev_rows[printer_id][i])
                        difference = current_value - prev_row_value
                        # Append current value if it's less than previous row's value
                        if current_value < prev_row_value:
                            diff_sheet[f"{col}{index}"] = current_value
                        else:
                            diff_sheet[f"{col}{index}"] = difference
                    except (TypeError, ValueError):
                        continue
            # Store the current row as the previous row for the next iteration
            prev_rows[printer_id] = row
        
        # Save the workbook
        wb.save(excel_file)
        print("Differences calculated and saved in Printer Daily Usage sheet.")
    
    except Exception as e:
        print(f"Error occurred: {e}")

def create_printer_usage_table(excel_file):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        
        # Select the worksheet containing printer daily usage data
        daily_usage_sheet = wb["Printer Daily Usage"]
        
        # Create a new worksheet for the printer usage table
        usage_sheet = wb.create_sheet(title="Printer Usage")
        
        # Add headers to the usage sheet
        usage_sheet.append(["Printer ID", "Page ID", "Page Count"])
        
        # Iterate over rows starting from the second row (assuming headers are in the first row)
        for row in daily_usage_sheet.iter_rows(min_row=2, values_only=True):
            printer_id, _, _, _, _, a4_page, a5_page = row
            
            # Check if both A4 page and A5 page are not empty
            if a4_page is not None and a5_page is not None:
                # Append data to the usage sheet with page ID 1 for A4 and page ID 2 for A5
                usage_sheet.append([printer_id, 1, a4_page])
                usage_sheet.append([printer_id, 2, a5_page])
        
        # Save the workbook
        wb.save(excel_file)
        print("Printer usage table created in the Printer Usage sheet.")
    
    except Exception as e:
        print(f"Error occurred: {e}")


# Example usage
excel_file = "Printer_Metrics.xlsx"
sheet_name = "printers"

calculate_difference(excel_file, sheet_name)
create_printer_usage_table(excel_file)
